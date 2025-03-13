import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path

from openpyxl.styles.builtins import total

# Ensure that the PDFs directory exists, if not, create it
output_dir = Path("PDFs")
output_dir.mkdir(parents=True, exist_ok=True)  # Create the directory if it doesn't exist

# Get all file paths with .xlsx extension in the invoices folder
filepaths = glob.glob("invoices/*.xlsx")

# Loop through each file path and read the corresponding Excel file
for filepath in filepaths:
    try:
        # Read the sheet named "Sheet 1"
        df = pd.read_excel(filepath, sheet_name="Sheet 1")  # Read the sheet named "Sheet 1"

        # Initialize PDF
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()

        # Extract filename and invoice number from the filepath
        filename = Path(filepath).stem  # Get the filename without the extension
        invoice_nr,date = filename.split("-")  # Assuming the invoice number is the first part of the filename

        # Set the font for the invoice number
        pdf.set_font(family="Times", size=16, style="B")

        # Add the invoice number to the PDF
        pdf.cell(w=50, h=8, txt=f"Invoice nr.{invoice_nr}", ln=1)


        pdf.set_font(family="Times", size=16, style="B")
        pdf.cell(w=50, h=8, txt=f"Date: {date}",ln=1)

        columns = list(df.columns)
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30,h=8,txt= columns[0],border=1)
        pdf.cell(w=70,h=8,txt= columns[1],border=1)
        pdf.cell(w=30,h=8,txt= columns[2],border=1)
        pdf.cell(w=30,h=8,txt= columns[3],border=1)
        pdf.cell(w=30,h=8,txt= columns[4],border=1,ln=1)

        for index,row in df.iterrows():
            pdf.set_font(family="Times", size=10)
            pdf.set_text_color(80, 80, 80)
            pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
            pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
            pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)
            pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
            pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

        total_sum = df["total_price"].sum()
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30, h=8, txt="", border=1)
        pdf.cell(w=70, h=8, txt="", border=1)
        pdf.cell(w=30, h=8, txt="", border=1)
        pdf.cell(w=30, h=8, txt="", border=1)
        pdf.cell(w=30, h=8, txt=str(total_sum), border=1, ln=1)


        # Save the PDF to the output directory
        pdf.output(output_dir / f"{filename}.pdf")  # Save PDF in the PDFs folder

    except Exception as e:
        print(f"Error processing file {filepath}: {e}")
